from __future__ import annotations

from pathlib import Path

import openpyxl
from docx import Document
from pypdf import PdfReader


SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xls"}


def load_document_file(filepath: str | Path) -> list[dict]:
    path = Path(filepath)
    suffix = path.suffix.lower()

    if suffix == ".docx":
        return _single_document(path, _load_docx_text(path), "docx")

    if suffix == ".pdf":
        return _single_document(path, _load_pdf_text(path), "pdf")

    if suffix in {".xlsx", ".xls"}:
        return _load_workbook_docs(path)

    raise ValueError(f"Unsupported file type: {suffix}")


def load_documents(documents_dir: str | Path) -> tuple[list[dict], list[str]]:
    docs = []
    warnings = []
    directory = Path(documents_dir)

    for path in sorted(directory.iterdir()):
        if not path.is_file():
            continue
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            warnings.append(f"Skipped unsupported file: {path.name}")
            continue
        try:
            docs.extend(load_document_file(path))
        except Exception as exc:
            warnings.append(f"Could not load {path.name}: {exc}")

    return docs, warnings


def _single_document(path: Path, text: str, doc_type: str) -> list[dict]:
    if not text.strip():
        return []
    return [{"filename": path.name, "text": _source_prefixed_text(path.name, text), "type": doc_type}]


def _load_docx_text(path: Path) -> str:
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _load_pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join(page_text for page in reader.pages if (page_text := page.extract_text()))


def _load_workbook_docs(path: Path) -> list[dict]:
    workbook = openpyxl.load_workbook(path, data_only=True)
    docs = []
    for sheet_name in workbook.sheetnames:
        sheet_text = _sheet_to_text(workbook[sheet_name], sheet_name)
        if sheet_text.strip():
            docs.append({
                "filename": f"{path.name}-{sheet_name}",
                "text": _source_prefixed_text(path.name, sheet_text, sheet_name=sheet_name),
                "type": "xlsx",
            })
    return docs


def _sheet_to_text(sheet, sheet_name: str) -> str:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""

    headers = [str(cell) if cell is not None else "" for cell in rows[0]]
    text_parts = [f"Sheet: {sheet_name}"]
    for row in rows[1:]:
        row_text = _format_row(headers, row)
        if row_text:
            text_parts.append(row_text)
    return "\n".join(text_parts)


def _format_row(headers: list[str], row: tuple) -> str:
    return " | ".join(f"{header}: {cell}" for header, cell in zip(headers, row) if cell is not None)


def _source_prefixed_text(filename: str, text: str, sheet_name: str | None = None) -> str:
    parts = [f"Document filename: {filename}"]
    if sheet_name:
        parts.append(f"Sheet: {sheet_name}")
    return "\n".join(parts) + f"\n\n{text}"
